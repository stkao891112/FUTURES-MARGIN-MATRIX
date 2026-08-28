import os
import re
import sys
import time
import json
import concurrent.futures
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except Exception:
        pass

from binance import crawl_binance_leverage_margin
from bybit import crawl_bybit_multi_coins_to_excel
from bitget import crawl_bitget_position_tiers
from okx import crawl_okx_position_tiers
from mexc import crawl_mexc_position_tiers
from bingx import crawl_bingx_position_tiers
from pionex import crawl_pionex_position_tiers

def extract_tier_number(val):
    """
    從檔位字串提取數字做為排序依據（例如 "檔位 1" -> 1, "2" -> 2）
    """
    if pd.isna(val):
        return 999999
    val_str = str(val).strip()
    match = re.search(r'\d+', val_str)
    return int(match.group()) if match else 999999

def clean_range_upper_bound(val_str):
    """
    清理範圍字串，僅保留上限數值（例如 "0~200,000" -> "200,000", "200,000~1,000,000" -> "1,000,000"）
    """
    if not isinstance(val_str, str):
        return val_str
    val_str = str(val_str).strip()
    for sep in ['~', '～']:
        if sep in val_str:
            parts = val_str.split(sep)
            if len(parts) > 1 and parts[-1].strip():
                return parts[-1].strip()
    return val_str

def normalize_dataframe_columns(df):
    """
    統一四大交易所的欄位名稱，以 Binance / 7大標準結構為基準對齊：
    1. 交易所
    2. 檔位
    3. 單位 (OKX視搜尋幣種，其餘固定USDT)
    4. 倉位分級
    5. 最大槓桿
    6. 維持保證金率
    7. 維持金額(USDT)
    """
    if df.empty:
        return df

    column_mapping = {}
    for col in df.columns:
        c_str = str(col).strip()
        if c_str == "交易所":
            column_mapping[col] = "交易所"
        elif c_str in ["檔位", "等級", "欄位_1"]:
            column_mapping[col] = "檔位"
        elif c_str == "單位":
            column_mapping[col] = "單位"
        elif any(k in c_str for k in ["倉位", "風險限額", "價值", "張數", "名義"]) or c_str == "欄位_2":
            column_mapping[col] = "倉位分級"
        elif "槓桿" in c_str or c_str == "欄位_3":
            column_mapping[col] = "最大槓桿"
        elif any(k in c_str for k in ["金額", "扣減額"]) or c_str == "欄位_5":
            column_mapping[col] = "維持金額(USDT)"
        elif any(k in c_str for k in ["保證金"]) or c_str == "欄位_4":
            column_mapping[col] = "維持保證金率"

    renamed_df = df.rename(columns=column_mapping)

    # 若缺乏「單位」欄位，預設補上 USDT
    if "單位" not in renamed_df.columns:
        renamed_df["單位"] = "USDT"

    # 針對倉位價值/範圍欄位自動萃取上限數值 (如 0~200,000 -> 200,000)
    if "倉位分級" in renamed_df.columns:
        renamed_df["倉位分級"] = renamed_df["倉位分級"].apply(clean_range_upper_bound)

    standard_headers = ["交易所", "檔位", "單位", "倉位分級", "最大槓桿", "維持保證金率", "維持金額(USDT)"]
    final_cols = [c for c in standard_headers if c in renamed_df.columns]
    extra_cols = [c for c in renamed_df.columns if c not in final_cols]
    return renamed_df[final_cols + extra_cols]

def merge_and_sort_exchanges_data(coin_dfs):
    """
    合併多個交易所的 DataFrame，將欄位統一對齊，
    並依據「交易所」名稱與「檔位」進行排序。
    """
    if not coin_dfs:
        return pd.DataFrame()

    normalized_dfs = [normalize_dataframe_columns(df) for df in coin_dfs if not df.empty]
    if not normalized_dfs:
        return pd.DataFrame()

    combined_df = pd.concat(normalized_dfs, ignore_index=True)

    # 確保「交易所」欄位在第 1 欄 (A 欄)
    if "交易所" in combined_df.columns:
        cols = ["交易所"] + [c for c in combined_df.columns if c != "交易所"]
        combined_df = combined_df[cols]

    # 依「交易所」與「檔位」排序
    if "檔位" in combined_df.columns and "交易所" in combined_df.columns:
        combined_df["__tier_sort"] = combined_df["檔位"].apply(extract_tier_number)
        combined_df.sort_values(by=["交易所", "__tier_sort"], ascending=[True, True], inplace=True)
        combined_df.drop(columns=["__tier_sort"], inplace=True)
    elif "交易所" in combined_df.columns:
        combined_df.sort_values(by=["交易所"], ascending=[True], inplace=True)

    return combined_df

def apply_excel_styles_and_colors(ws):
    """
    為 Excel 工作表套用色彩標籤與專業樣式：
    - Binance: 柔和黃色 (#FFF2CC)
    - Bitget: 柔水藍色 (#E0F2FE)
    - Bybit: 柔和橘色 (#FFE0B2)
    - OKX: 柔和淡紫 / 灰藍 (#E8EAF6)
    - 標題列: 深灰藍色 (#2C3E50) 與白色粗體
    """
    fill_binance = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 幣安黃
    fill_bitget = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")   # Bitget 水藍
    fill_bybit = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")    # Bybit 橘
    fill_okx = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")      # OKX 淡紫
    fill_mexc = PatternFill(start_color="E0F2F1", end_color="E0F2F1", fill_type="solid")     # MEXC 青綠
    fill_bingx = PatternFill(start_color="EBF3FF", end_color="EBF3FF", fill_type="solid")    # BingX 藍
    fill_pionex = PatternFill(start_color="FFF0E5", end_color="FFF0E5", fill_type="solid")   # Pionex 橘
    fill_default = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    fill_header = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    font_header = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="微軟正黑體", size=10)

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    # 1. 樣式化標題列 (Row 1)
    for cell in ws[1]:
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center
        cell.border = thin_border

    # 2. 逐行設定資料儲存格背景色與邊框
    for row in range(2, ws.max_row + 1):
        ex_val = str(ws.cell(row=row, column=1).value or "").strip()

        if "Binance" in ex_val:
            row_fill = fill_binance
        elif "Bitget" in ex_val:
            row_fill = fill_bitget
        elif "Bybit" in ex_val:
            row_fill = fill_bybit
        elif "OKX" in ex_val:
            row_fill = fill_okx
        elif "MEXC" in ex_val:
            row_fill = fill_mexc
        elif "BingX" in ex_val:
            row_fill = fill_bingx
        elif "Pionex" in ex_val:
            row_fill = fill_pionex
        else:
            row_fill = fill_default

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = row_fill
            cell.font = font_data
            cell.border = thin_border
            
            if col in [1, 2, 3]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    # 3. 自動調整欄寬
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_str = str(cell.value or '')
            display_len = sum(2 if ord(c) > 127 else 1 for c in val_str)
            max_len = max(max_len, display_len)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def write_crawl_status(ex_name, coin_list, status_type, rows_info=0):
    status_file = os.path.join(BASE_DIR, "data", "crawl_status.json")
    status_data = {}
    if os.path.exists(status_file):
        try:
            with open(status_file, "r", encoding="utf-8") as f:
                status_data = json.load(f)
        except Exception:
            pass

    for coin in coin_list:
        if coin not in status_data:
            status_data[coin] = {}
        
        # 支援針對各幣種精準寫入個別檔位數
        count = rows_info.get(coin, 0) if isinstance(rows_info, dict) else rows_info
        status_data[coin][ex_name] = {
            "status": status_type,
            "count": count,
            "updated_at": time.strftime("%H:%M:%S")
        }

    try:
        with open(status_file, "w", encoding="utf-8") as f:
            json.dump(status_data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def run_single_exchange(args):
    """
    線程池執行的單一交易所抓取封裝函式
    """
    name, crawl_func, coins = args
    print(f"\n⚡ [並行任務啟動] 開始抓取 [{name}] 數據...")
    write_crawl_status(name, coins, "running")
    start_time = time.time()
    try:
        res = crawl_func(coins, save_excel=False)
        elapsed = time.time() - start_time
        counts_by_coin = {c: len(res[c]) for c in coins if c in (res or {}) and hasattr(res[c], '__len__')}
        write_crawl_status(name, coins, "success", counts_by_coin)
        return name, res or {}, f"✅ 成功 (耗時 {elapsed:.1f} 秒)"
    except Exception as e:
        elapsed = time.time() - start_time
        write_crawl_status(name, coins, "failed", 0)
        return name, {}, f"❌ 失敗: {e} (耗時 {elapsed:.1f} 秒)"

def audit_and_verify_all_data(all_results, coins, task_func_map=None, max_retries=1):
    """
    全交易所全幣種數據完整性檢核機制：
    1. 逐一檢查 [幣種 x 交易所] 矩陣，檢驗是否有任一交易所任一幣種缺失或 0 檔位
    2. 自動啟動針對缺失幣種之「專項重試機制」
    3. 生成並列印專業的數據完整性檢核矩陣表格
    4. 輸出檢核報告至 data/audit_report.json
    """
    exchanges = ["Binance (幣安)", "Bybit", "Bitget", "OKX", "MEXC", "BingX", "Pionex"]
    short_names = {
        "Binance (幣安)": "Binance",
        "Bybit": "Bybit",
        "Bitget": "Bitget",
        "OKX": "OKX",
        "MEXC": "MEXC",
        "BingX": "BingX",
        "Pionex": "Pionex"
    }

    # 1. 第一輪檢查缺失清單
    missing_by_ex = {}
    for ex in exchanges:
        res_dict = all_results.get(ex, {})
        for coin in coins:
            df = res_dict.get(coin)
            if df is None or (hasattr(df, 'empty') and df.empty) or len(df) == 0:
                missing_by_ex.setdefault(ex, []).append(coin)

    # 2. 自動針對缺失項目進行專項重試 (Auto-Retry)
    if missing_by_ex and task_func_map and max_retries > 0:
        print("\n" + "=" * 65)
        print(" 🔁 觸發數據完整性自動補救重試機制 (Auto-Retry for Missing Coins)")
        print("=" * 65)
        for ex, miss_coins in missing_by_ex.items():
            crawl_func = task_func_map.get(ex)
            if crawl_func:
                print(f"🔄 正在為 [{ex}] 補抓缺失幣種: {miss_coins}...")
                try:
                    retry_res = crawl_func(miss_coins, save_excel=False)
                    if retry_res:
                        if ex not in all_results:
                            all_results[ex] = {}
                        for mc in miss_coins:
                            if mc in retry_res and not retry_res[mc].empty:
                                all_results[ex][mc] = retry_res[mc]
                                print(f"  ✅ [{ex}] -> [{mc}]: 補抓成功 (共 {len(retry_res[mc])} 檔)！")
                            else:
                                print(f"  ❌ [{ex}] -> [{mc}]: 補抓依然無數據")
                except Exception as e:
                    print(f"  ❌ 重試 [{ex}] 時發生異常: {e}")

    # 3. 建立最終檢核矩陣與統計
    matrix = {}
    total_checkpoints = len(coins) * len(exchanges)
    passed_checkpoints = 0
    missing_details = []

    for coin in coins:
        matrix[coin] = {}
        c_passed = 0
        for ex in exchanges:
            res_dict = all_results.get(ex, {})
            df = res_dict.get(coin)
            count = len(df) if (df is not None and hasattr(df, '__len__')) else 0
            if count > 0:
                matrix[coin][ex] = {"status": "ok", "count": count}
                passed_checkpoints += 1
                c_passed += 1
            else:
                matrix[coin][ex] = {"status": "missing", "count": 0}
                missing_details.append(f"{ex} - {coin}")
        matrix[coin]["_completeness"] = f"{c_passed}/{len(exchanges)}"

    coverage_rate = (passed_checkpoints / total_checkpoints) * 100 if total_checkpoints > 0 else 0

    # 4. 輸出視覺化檢核矩陣
    col_w = 11
    headers = [f"{short_names[ex]:<{col_w}}" for ex in exchanges]
    header_str = "幣種          " + " ".join(headers) + " 完整率"
    div_line = "-" * len(header_str)

    print("\n" + "=" * len(header_str))
    print(" 📋 全交易所幣種數據完整性檢核矩陣 (Integrity Audit Matrix)")
    print("=" * len(header_str))
    print(header_str)
    print(div_line)

    for coin in coins:
        row_cells = []
        for ex in exchanges:
            info = matrix[coin][ex]
            if info["status"] == "ok":
                cell = f"✅ {info['count']}檔"
            else:
                cell = "❌ 缺失"
            row_cells.append(f"{cell:<{col_w}}")
        print(f"{coin:<14}" + " ".join(row_cells) + f" {matrix[coin]['_completeness']}")

    print(div_line)
    print(f"📌 檢核點總數: {total_checkpoints} 個 ({len(coins)} 幣種 x {len(exchanges)} 交易所)")
    print(f"✅ 通過數量:   {passed_checkpoints} / {total_checkpoints} ({coverage_rate:.1f}%)")
    if missing_details:
        print(f"❌ 缺失項目 ({len(missing_details)} 項):")
        for m in missing_details:
            print(f"   • {m}")
        print("🚨【數據完整性警告】部分交易所與幣種數據缺失，請確認標的是否在該交易所上架或檢查網路。")
    else:
        print("🎉【100% 數據完整性驗證通過】所有交易所之所有幣種均已成功獲取真實階梯檔位！")
    print("=" * len(header_str))

    # 5. 輸出報告 JSON
    audit_report = {
        "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
        "total_checkpoints": total_checkpoints,
        "passed_checkpoints": passed_checkpoints,
        "coverage_rate": round(coverage_rate, 2),
        "is_all_passed": (passed_checkpoints == total_checkpoints),
        "missing_count": len(missing_details),
        "missing_details": missing_details,
        "matrix": matrix
    }

    try:
        report_file = os.path.join(BASE_DIR, "data", "audit_report.json")
        with open(report_file, "w", encoding="utf-8") as f:
            json.dump(audit_report, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

    return audit_report

def verify_existing_tiers_data():
    """
    直接從現有 data/tiers.json 進行快速完整性審查
    """
    tiers_path = os.path.join(BASE_DIR, "data", "tiers.json")
    if not os.path.exists(tiers_path):
        print(f"❌ 未找到 [{tiers_path}]，請先執行爬蟲！")
        return
    with open(tiers_path, "r", encoding="utf-8") as f:
        tiers_data = json.load(f)
    
    coins = [k for k in tiers_data.keys() if not k.startswith("_")]
    all_results = {}
    ex_name_map = {
        "Binance": "Binance (幣安)",
        "Bybit": "Bybit",
        "Bitget": "Bitget",
        "OKX": "OKX",
        "MEXC": "MEXC",
        "BingX": "BingX",
        "Pionex": "Pionex"
    }
    for ex_key, full_name in ex_name_map.items():
        all_results[full_name] = {}
        for c in coins:
            c_data = tiers_data[c].get(ex_key, [])
            all_results[full_name][c] = pd.DataFrame(c_data)
    
    audit_and_verify_all_data(all_results, coins)

def main():
    print("=" * 60)
    print(" 🚀 跨交易所合約保證金數據自動化爬蟲（7家並行版 + 數據完整性檢核機制） ")
    print("=" * 60)
    
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    coins_path = os.path.join(BASE_DIR, "data", "coins.json")

    # 支援純檢核模式 (python main.py --verify 或 -v)
    if any(arg in ['--verify', '--check', '-v'] for arg in sys.argv[1:]):
        print("🔍 啟動獨立快速數據審查模式...")
        verify_existing_tiers_data()
        return

    # 支援增量爬取 (命令列傳入 --coin DOGEUSDT 或位置參數 DOGEUSDT)
    coins_arg = []
    if len(sys.argv) > 1:
        args_list = sys.argv[1:]
        for i, arg in enumerate(args_list):
            if arg in ['--coin', '--coins', '-c'] and i + 1 < len(args_list):
                raw = args_list[i + 1]
                coins_arg.extend([c.strip().upper() for c in raw.split(',') if c.strip()])
            elif not arg.startswith('-'):
                coins_arg.extend([c.strip().upper() for c in arg.split(',') if c.strip()])

        clean_coins = []
        for c in coins_arg:
            if not c.endswith('USDT') and not c.endswith('USD'):
                c += 'USDT'
            if c not in clean_coins:
                clean_coins.append(c)
        if clean_coins:
            coins = clean_coins

    if not 'coins' in locals() or not coins:
        if os.path.exists(coins_path):
            try:
                with open(coins_path, "r", encoding="utf-8") as f:
                    coins = json.load(f)
            except Exception as e:
                print("❌ 載入 coins.json 失敗:", e)
                coins = ["BTCUSDT", "ETHUSDT", "SOLUSDT"]
        else:
            coins = ["BTCUSDT", "ETHUSDT", "SOLUSDT"]

    print(f"\n📌 本次任務將為以下 {len(coins)} 個幣種抓取各大交易所數據: {coins}")
    
    tasks = [
        ("Binance (幣安)", crawl_binance_leverage_margin, coins),
        ("Bybit", crawl_bybit_multi_coins_to_excel, coins),
        ("Bitget", crawl_bitget_position_tiers, coins),
        ("OKX", crawl_okx_position_tiers, coins),
        ("MEXC", crawl_mexc_position_tiers, coins),
        ("BingX", crawl_bingx_position_tiers, coins),
        ("Pionex", crawl_pionex_position_tiers, coins),
    ]

    all_results = {}
    summary_results = {}

    print("\n" + "=" * 60)
    print(f" ⚡ 正在同時 (並行) 啟動 {len(tasks)} 家交易所 Playwright 瀏覽器...")
    print("=" * 60)

    # 使用 ThreadPoolExecutor 同時啟動 7 家爬蟲
    start_total_time = time.time()
    with concurrent.futures.ThreadPoolExecutor(max_workers=len(tasks)) as executor:
        futures = [executor.submit(run_single_exchange, task) for task in tasks]
        for future in concurrent.futures.as_completed(futures):
            ex_name, res, status = future.result()
            all_results[ex_name] = res
            summary_results[ex_name] = status

    total_elapsed = time.time() - start_total_time

    # 執行最終數據完整性檢核機制 (含自動補救重試)
    task_func_map = {name: func for name, func, _ in tasks}
    audit_and_verify_all_data(all_results, coins, task_func_map=task_func_map, max_retries=1)

    # 進行同一幣種跨交易所數據整合
    data_dir = os.path.join(BASE_DIR, "data")
    if not os.path.exists(data_dir):
        os.makedirs(data_dir)

    combined_excel_filename = os.path.join(data_dir, "all_exchanges_futures_margin.xlsx")
    print("\n" + "=" * 60)
    print(" 🔀 正在進行同幣種跨交易所數據合併與排序...")
    print("=" * 60)

    excel_kwargs = {'engine': 'openpyxl'}
    if os.path.exists(combined_excel_filename):
        excel_kwargs['mode'] = 'a'
        excel_kwargs['if_sheet_exists'] = 'replace'

    with pd.ExcelWriter(combined_excel_filename, **excel_kwargs) as writer:
        for coin in coins:
            coin_dfs = []
            for ex_name, res_dict in all_results.items():
                if coin in res_dict and not res_dict[coin].empty:
                    coin_dfs.append(res_dict[coin])

            if coin_dfs:
                merged_df = merge_and_sort_exchanges_data(coin_dfs)
                sheet_name = coin[:31]
                merged_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 套用背景色彩與專業格式
                ws = writer.sheets[sheet_name]
                apply_excel_styles_and_colors(ws)
                print(f" 💾 已成功將 [{coin}] 數據寫入工作表: [{sheet_name}] (共 {len(merged_df)} 列)")
            else:
                print(f" ⚠️ [{coin}] 無有效數據可供合併。")

    # 印出執行總結報告
    print("\n" + "=" * 60)
    print(" 📊 總結報告 (Task Summary)")
    print("=" * 60)
    for ex_name, status in summary_results.items():
        print(f" • {ex_name:<20}: {status}")
    print(f"\n⏱️  並行爬蟲總耗時: {total_elapsed:.1f} 秒")

    if os.path.exists(combined_excel_filename):
        size_kb = os.path.getsize(combined_excel_filename) / 1024
        print(f"\n🎉 整合完成！同幣種跨交易所匯總 Excel 已儲存至:")
        print(f" 📂 {os.path.abspath(combined_excel_filename)} (大小: {size_kb:.1f} KB)")
        
        # 自動轉出 JSON 供網頁版計算機實時使用
        try:
            from export_json import export_tiers_to_json
            export_tiers_to_json()
        except Exception:
            pass

if __name__ == "__main__":
    main()
